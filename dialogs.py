import os
import shutil
import time

from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QFileDialog,
    QMessageBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QProgressBar, QAbstractItemView,
    QCheckBox, QWidget, QFrame
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont


# =========================
# PRODUCT ADD / EDIT DIALOG
# =========================
class ProductDialog(QDialog):
    def __init__(self, parent=None, product_data=None):
        super().__init__(parent)
        self.product_data = product_data
        self.current_image_path = product_data[4] if product_data else ""

        title = "Edit Produk" if product_data else "Tambah Produk Baru"
        self.setWindowTitle(title)
        self.setFixedSize(400, 480)
        self.setStyleSheet("background-color: white; border-radius: 8px;")

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)

        lbl_title = QLabel("✏️ Edit Detail Produk" if self.product_data else "➕ Tambah Produk")
        lbl_title.setStyleSheet("font-size: 18px; font-weight: bold; color: #31353b;")
        layout.addWidget(lbl_title)

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignLeft)
        form_layout.setVerticalSpacing(15)

        self.name = QLineEdit()
        self.name.setPlaceholderText("Masukkan nama produk...")

        self.stock = QLineEdit()
        self.stock.setPlaceholderText("0")

        self.price = QLineEdit()
        self.price.setPlaceholderText("0")

        self.btn_img = QPushButton("📸 Pilih Gambar")
        self.btn_img.setObjectName("btnSecondary")
        self.btn_img.clicked.connect(self.select_image)

        self.lbl_img_status = QLabel(
            "Tidak ada gambar baru dipilih" if self.product_data else "Belum ada gambar dipilih"
        )
        self.lbl_img_status.setStyleSheet("color: #6d7588; font-size: 11px; font-style: italic;")
        self.lbl_img_status.setWordWrap(True)

        if self.product_data:
            self.name.setText(self.product_data[1])
            self.stock.setText(str(self.product_data[2]))
            self.price.setText(str(int(self.product_data[3])))
            if self.current_image_path:
                self.lbl_img_status.setText(f"Gambar saat ini: {os.path.basename(self.current_image_path)}")

        lbl_style = "font-weight: 600; color: #31353b; font-size: 13px;"
        l_name = QLabel("Nama Produk"); l_name.setStyleSheet(lbl_style)
        l_stock = QLabel("Stok Barang"); l_stock.setStyleSheet(lbl_style)
        l_price = QLabel("Harga (Rp)"); l_price.setStyleSheet(lbl_style)
        l_img = QLabel("Cover Produk"); l_img.setStyleSheet(lbl_style)

        form_layout.addRow(l_name, self.name)
        form_layout.addRow(l_stock, self.stock)
        form_layout.addRow(l_price, self.price)
        form_layout.addRow(l_img, self.btn_img)
        form_layout.addRow("", self.lbl_img_status)

        layout.addLayout(form_layout)
        layout.addStretch()

        btn_layout = QHBoxLayout()
        self.btn_cancel = QPushButton("Batal")
        self.btn_cancel.setObjectName("btnSecondary")
        self.btn_cancel.clicked.connect(self.reject)

        self.btn_save = QPushButton("Simpan")
        self.btn_save.setObjectName("btnPrimary")
        self.btn_save.clicked.connect(self.save_data)

        btn_layout.addWidget(self.btn_cancel)
        btn_layout.addWidget(self.btn_save)
        layout.addLayout(btn_layout)

    def select_image(self):
        file, _ = QFileDialog.getOpenFileName(self, "Pilih Gambar", "", "Images (*.png *.jpg *.jpeg)")
        if file:
            self.current_image_path = file
            self.lbl_img_status.setText(f"Terpilih: {os.path.basename(file)}")
            self.lbl_img_status.setStyleSheet("color: #42b549; font-size: 11px; font-weight: bold;")

    def save_data(self):
        n = self.name.text().strip()
        s = self.stock.text().strip()
        p = self.price.text().strip()

        if not n or not s or not p:
            QMessageBox.warning(self, "Peringatan", "Semua kolom teks harus diisi!")
            return

        try:
            stock = int(s)
            price = float(p)
        except ValueError:
            QMessageBox.critical(self, "Error", "Stok dan Harga harus berupa angka!")
            return

        final_img_path = self.current_image_path
        if (self.current_image_path
                and os.path.exists(self.current_image_path)
                and not self.current_image_path.startswith("images/")):
            ext = os.path.splitext(self.current_image_path)[1]
            if not os.path.exists("images"):
                os.makedirs("images")
            new = f"images/item_{int(time.time())}{ext}"
            shutil.copy(self.current_image_path, new)
            final_img_path = new

        self.result_data = (n, stock, price, final_img_path)
        self.accept()


# =========================
# EXCEL IMPORT WORKER (Background Thread)
# =========================
class ExcelImportWorker(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(list, list)   # (valid_rows, skipped_rows)
    error = pyqtSignal(str)

    # Mapping kolom Excel -> indeks (berdasarkan Dataset.xlsx)
    COL_KODE = 0
    COL_BERAT = 1
    COL_COLUMN1 = 2
    COL_NAME = 3
    COL_DESC = 4
    COL_HPP = 5
    COL_HARGA_GROSIR = 6
    COL_HARGA_TOKO = 7
    COL_JUMLAH_MASUK = 8
    COL_ISI_BAL = 9
    COL_TGL = 10
    COL_STOCK_AKHIR = 11
    COL_HARGA_SHOPEE = 12
    COL_AWAL = 13
    COL_MASUK = 14
    COL_KELUAR = 15
    COL_AKHIR = 16
    COL_ISI_BAL2 = 17

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path

    def safe_float(self, value):
        if value is None: return 0.0
        # Jika sudah berupa angka asli, langsung kembalikan dan bulatkan (hindari error floating point)
        if isinstance(value, (int, float)):
            return round(float(value), 2)
            
        val_str = str(value).strip().lower()
        if val_str in ["xxx", "-", "na", "n/a", ""]: return 0.0
        try:
            return round(float(val_str.replace(".", "").replace(",", "")), 2)
        except ValueError:
            return 0.0

    def safe_int(self, value):
        if value is None: return 0
        # Jika sudah berupa angka asli, langsung jadikan integer yang dibulatkan
        if isinstance(value, (int, float)):
            return int(round(value))
            
        val_str = str(value).strip().lower()
        if val_str in ["xxx", "-", "na", "n/a", ""]: return 0
        try:
            return int(round(float(val_str.replace(".", "").replace(",", ""))))
        except ValueError:
            return 0

    def run(self):
        try:
            from openpyxl import load_workbook
            # Menggunakan data_only=True agar membaca hasil formula, bukan text rumusnya
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active

            total_rows = ws.max_row if ws.max_row else 1000

            valid_rows = []
            skipped_rows = []

            # MENCEGAH CRASH OPENPYXL:
            for i, row_cells in enumerate(ws.iter_rows(values_only=False)):
                if total_rows > 0:
                    self.progress.emit(min(99, int((i + 1) / total_rows * 100)))

                row = []
                for cell in row_cells:
                    try:
                        val = cell.value
                    except Exception:
                        val = "" # Bypass aman jika openpyxl internal error
                    row.append(val)

                # Fix: Bounds checking, skip entirely empty rows safely
                if not row or len(row) <= self.COL_NAME or row[self.COL_NAME] is None:
                    continue
                # Lewati baris header literal
                if str(row[self.COL_NAME]).strip().upper() in ("NAMA BARANG", ""):
                    continue

                name  = str(row[self.COL_NAME]).strip()
                
                # --- PARSING HARGA TOKO AMAN DARI FLOAT ERROR ---
                price_raw = row[self.COL_HARGA_TOKO] if len(row) > self.COL_HARGA_TOKO else 0
                
                if isinstance(price_raw, (int, float)):
                    price = round(float(price_raw), 2)
                else:
                    price_str = str(price_raw).strip().lower() if price_raw is not None else ""
                    if price_str in ["xxx", "-", "na", "n/a", ""]:
                        price = 0.0
                    else:
                        try:
                            price = round(float(price_str.replace(".", "").replace(",", "")), 2)
                        except ValueError:
                            skipped_rows.append({
                                "nama_barang": name,
                                "reason": f"Harga tidak valid: {price_raw}"
                            })
                            continue
                        
                # --- PARSING STOK AMAN DARI FLOAT ERROR ---
                stock_raw = row[self.COL_JUMLAH_MASUK] if len(row) > self.COL_JUMLAH_MASUK else 0
                
                if isinstance(stock_raw, (int, float)):
                    stock = int(round(stock_raw))
                else:
                    stock_str = str(stock_raw).strip().lower() if stock_raw is not None else ""
                    if stock_str in ["xxx", "-", "na", "n/a", ""]:
                        stock = 0
                    else:
                        try:
                            stock = int(round(float(stock_str.replace(".", "").replace(",", ""))))
                        except ValueError:
                            skipped_rows.append({
                                "nama_barang": name,
                                "reason": f"Stok tidak valid: {stock_raw}"
                            })
                            continue

                valid_rows.append({
                    "kode_barang": str(row[self.COL_KODE]).strip() if len(row) > self.COL_KODE and row[self.COL_KODE] else "",
                    "berat_packing": self.safe_float(row[self.COL_BERAT] if len(row) > self.COL_BERAT else 0),
                    "column1": str(row[self.COL_COLUMN1] or "") if len(row) > self.COL_COLUMN1 else "",
                    "nama_barang": name,
                    "desc": str(row[self.COL_DESC] or "") if len(row) > self.COL_DESC else "",
                    "hpp": self.safe_float(row[self.COL_HPP] if len(row) > self.COL_HPP else 0),
                    "harga_jual_grosir": self.safe_float(row[self.COL_HARGA_GROSIR] if len(row) > self.COL_HARGA_GROSIR else 0),
                    "harga_jual_toko": price,
                    "jumlah_masuk": stock,
                    "isi_per_bal": self.safe_int(row[self.COL_ISI_BAL] if len(row) > self.COL_ISI_BAL else 0),
                    "tgl": str(row[self.COL_TGL] or "") if len(row) > self.COL_TGL else "",
                    "stock_akhir": self.safe_int(row[self.COL_STOCK_AKHIR] if len(row) > self.COL_STOCK_AKHIR else 0),
                    "harga_shopee": self.safe_float(row[self.COL_HARGA_SHOPEE] if len(row) > self.COL_HARGA_SHOPEE else 0),
                    "awal": self.safe_int(row[self.COL_AWAL] if len(row) > self.COL_AWAL else 0),
                    "masuk": self.safe_int(row[self.COL_MASUK] if len(row) > self.COL_MASUK else 0),
                    "keluar": self.safe_int(row[self.COL_KELUAR] if len(row) > self.COL_KELUAR else 0),
                    "akhir": self.safe_int(row[self.COL_AKHIR] if len(row) > self.COL_AKHIR else 0),
                    "isi_per_bal2": self.safe_int(row[self.COL_ISI_BAL2] if len(row) > self.COL_ISI_BAL2 else 0),
                    "image_path": ""
                })

            wb.close()
            self.finished.emit(valid_rows, skipped_rows)

        except Exception as e:
            self.error.emit(str(e))


# =========================
# IMPORT EXCEL DIALOG
# =========================
class ImportExcelDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Import Produk dari Excel")
        self.setMinimumSize(780, 580)
        self.setStyleSheet("background-color: white;")

        self.parsed_data = []      # hasil parsing valid
        self.skipped_data = []     # baris yang dilewati
        self.worker = None

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)

        # --- Title ---
        lbl_title = QLabel("📥 Import Produk dari File Excel")
        lbl_title.setStyleSheet("font-size: 18px; font-weight: bold; color: #31353b;")
        layout.addWidget(lbl_title)

        lbl_sub = QLabel(
            "Pilih file Excel (.xlsx). Kolom yang dibaca: "
            "<b>NAMA BARANG</b>, <b>HARGA JUAL TOKO</b>, <b>JUMLAH MASUK</b>."
        )
        lbl_sub.setStyleSheet("color: #6d7588; font-size: 12px;")
        lbl_sub.setWordWrap(True)
        layout.addWidget(lbl_sub)

        # --- File picker ---
        file_row = QHBoxLayout()
        self.lbl_file = QLabel("Belum ada file dipilih")
        self.lbl_file.setStyleSheet(
            "color: #6d7588; font-size: 12px; font-style: italic;"
            "border: 1px dashed #d6dce5; border-radius: 6px; padding: 8px 12px;"
        )
        self.lbl_file.setSizePolicy(self.lbl_file.sizePolicy().horizontalPolicy(),
                                    self.lbl_file.sizePolicy().verticalPolicy())

        self.btn_browse = QPushButton("📂 Pilih File Excel")
        self.btn_browse.setObjectName("btnSecondary")
        self.btn_browse.setFixedWidth(170)
        self.btn_browse.clicked.connect(self.browse_file)

        file_row.addWidget(self.lbl_file)
        file_row.addWidget(self.btn_browse)
        layout.addLayout(file_row)

        # --- Progress ---
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: none; background-color: #f0f0f0;
                border-radius: 6px; height: 10px; text-align: center;
            }
            QProgressBar::chunk { background-color: #42b549; border-radius: 6px; }
        """)
        layout.addWidget(self.progress_bar)

        # --- Status label ---
        self.lbl_status = QLabel("")
        self.lbl_status.setStyleSheet("color: #6d7588; font-size: 12px;")
        layout.addWidget(self.lbl_status)

        # --- Preview Table ---
        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(3)
        self.preview_table.setHorizontalHeaderLabels(["Nama Produk", "Stok", "Harga Jual Toko"])
        self.preview_table.verticalHeader().setVisible(False)
        self.preview_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.preview_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setShowGrid(False)
        self.preview_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.preview_table.setColumnWidth(1, 100)
        self.preview_table.setColumnWidth(2, 160)
        layout.addWidget(self.preview_table)

        # --- Duplicate handling option ---
        self.chk_skip_dup = QCheckBox("Lewati produk dengan nama yang sudah ada di database")
        self.chk_skip_dup.setChecked(True)
        self.chk_skip_dup.setStyleSheet("color: #31353b; font-size: 12px;")
        layout.addWidget(self.chk_skip_dup)

        # --- Buttons ---
        btn_row = QHBoxLayout()
        self.btn_cancel = QPushButton("Batal")
        self.btn_cancel.setObjectName("btnSecondary")
        self.btn_cancel.clicked.connect(self.reject)

        self.btn_import = QPushButton("✅ Import ke Database")
        self.btn_import.setObjectName("btnPrimary")
        self.btn_import.setEnabled(False)
        self.btn_import.clicked.connect(self.confirm_import)

        btn_row.addStretch()
        btn_row.addWidget(self.btn_cancel)
        btn_row.addWidget(self.btn_import)
        layout.addLayout(btn_row)

    # ---- File Browsing ----
    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Pilih File Excel", "", "Excel Files (*.xlsx *.xlsm)"
        )
        if not file_path:
            return

        self.lbl_file.setText(os.path.basename(file_path))
        self.lbl_file.setStyleSheet(
            "color: #31353b; font-size: 12px; font-weight: bold;"
            "border: 1px solid #42b549; border-radius: 6px; padding: 8px 12px;"
        )
        self.btn_import.setEnabled(False)
        self.preview_table.setRowCount(0)
        self.parsed_data = []
        self.start_parsing(file_path)

    def start_parsing(self, file_path):
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.lbl_status.setText("Membaca file Excel...")
        self.btn_browse.setEnabled(False)

        self.worker = ExcelImportWorker(file_path)
        self.worker.progress.connect(self.progress_bar.setValue)
        self.worker.finished.connect(self.on_parse_done)
        self.worker.error.connect(self.on_parse_error)
        self.worker.start()

    def on_parse_done(self, valid_rows, skipped_rows):
        self.parsed_data = valid_rows
        self.skipped_data = skipped_rows
        self.progress_bar.setVisible(False)
        self.btn_browse.setEnabled(True)

        # Populate preview table (Fix: Map to correct DB parsed keys)
        self.preview_table.setRowCount(0)
        for row in valid_rows:
            r = self.preview_table.rowCount()
            self.preview_table.insertRow(r)
            self.preview_table.setItem(r, 0, QTableWidgetItem(row["nama_barang"]))

            stock_item = QTableWidgetItem(f"{row['jumlah_masuk']} Pcs")
            stock_item.setTextAlignment(Qt.AlignCenter)
            self.preview_table.setItem(r, 1, stock_item)

            price_item = QTableWidgetItem(f"Rp {row['harga_jual_toko']:,.0f}".replace(",", "."))
            font = price_item.font()
            font.setBold(True)
            price_item.setFont(font)
            self.preview_table.setItem(r, 2, price_item)

        skipped_info = f"  |  {len(skipped_rows)} baris dilewati" if skipped_rows else ""
        self.lbl_status.setText(
            f"✅ Berhasil membaca <b>{len(valid_rows)}</b> produk{skipped_info}. "
            f"Periksa data di atas, lalu klik <b>Import ke Database</b>."
        )
        self.btn_import.setEnabled(len(valid_rows) > 0)

    def on_parse_error(self, error_msg):
        self.progress_bar.setVisible(False)
        self.btn_browse.setEnabled(True)
        self.lbl_status.setText(f"❌ Gagal membaca file: {error_msg}")
        QMessageBox.critical(self, "Error", f"Gagal membaca file Excel:\n{error_msg}")

    # ---- Confirm & Import ----
    def confirm_import(self):
        count = len(self.parsed_data)
        reply = QMessageBox.question(
            self, "Konfirmasi Import",
            f"Anda akan mengimpor <b>{count}</b> produk ke database.\nLanjutkan?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes:
            self.result_data = self.parsed_data
            self.result_skip_dup = self.chk_skip_dup.isChecked()
            self.accept()